import argparse
import os

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

PROJECT_HEADERS = [
    "Project_ID",
    "Owner_Legal_Name",
    "GC_Legal_Name",
    "Project_Address",
    "County",
    "Latitude",
    "Longitude",
    "Disturbed_Acres",
    "Pre_Runoff_C",
    "Post_Runoff_C",
    "Receiving_Water",
    "Segment_No",
    "Inspection_Freq",
    "NOI_Issued_Date",
    "Permit_No",
    "Map_file_name",
]

LOTS_HEADERS = ["Project_ID", "Lot_No", "Start_Date", "End_Date"]

FREQ_LIST = ["7-day", "14-day", "Other"]

COMMENTS = {
    "Disturbed_Acres": "Total disturbed acreage. Example: 5.0",
    "Segment_No": "TCEQ classified segment number",
}


def build_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"

    ws.append(PROJECT_HEADERS)

    # Freeze header
    ws.freeze_panes = "A2"

    # Data validation lists
    dv_county = DataValidation(type="list", formula1="=Lists!$A:$A", allow_blank=True)
    dv_freq = DataValidation(type="list", formula1='"{}"'.format(",".join(FREQ_LIST)), allow_blank=True)
    ws.add_data_validation(dv_county)
    ws.add_data_validation(dv_freq)

    for idx, header in enumerate(PROJECT_HEADERS, start=1):
        col_letter = get_column_letter(idx)
        cell = ws[f"{col_letter}1"]
        cell.font = Font(bold=True)
        ws.column_dimensions[col_letter].auto_size = True
        # Apply validations
        if header == "County":
            dv_county.add(f"{col_letter}2:{col_letter}1048576")
        elif header == "Inspection_Freq":
            dv_freq.add(f"{col_letter}2:{col_letter}1048576")
        # Number formats
        if header in ("Latitude", "Longitude"):
            ws.column_dimensions[col_letter].number_format = "0.00000"
        elif header == "Disturbed_Acres":
            ws.column_dimensions[col_letter].number_format = "0.0"
        elif header in ("Pre_Runoff_C", "Post_Runoff_C"):
            ws.column_dimensions[col_letter].number_format = "0.00"
        # Comments
        if header in COMMENTS:
            ws[f"{col_letter}1"].comment = Comment(COMMENTS[header], "")

    # Unlock data cells
    for idx in range(1, len(PROJECT_HEADERS) + 1):
        col_letter = get_column_letter(idx)
        for row in range(2, 102):
            ws[f"{col_letter}{row}"].protection = Protection(locked=False)

    # Secondary sheet Lots
    ws2 = wb.create_sheet("Lots")
    ws2.append(LOTS_HEADERS)
    ws2.freeze_panes = "A2"
    for idx, header in enumerate(LOTS_HEADERS, start=1):
        col_letter = get_column_letter(idx)
        ws2[f"{col_letter}1"].font = Font(bold=True)
        ws2.column_dimensions[col_letter].auto_size = True
        for row in range(2, 102):
            ws2[f"{col_letter}{row}"].protection = Protection(locked=False)

    # Hidden sheet for County list
    lists = wb.create_sheet("Lists")
    lists.sheet_state = "hidden"
    lists.append(["County"])

    return wb


def save_workbook(wb, path, force=False):
    if os.path.exists(path) and not force:
        raise FileExistsError(f"{path} already exists. Use --force to overwrite")
    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description="Generate variable template")
    parser.add_argument("--out", default="templates/variable_template.xlsx")
    parser.add_argument("--force", action="store_true")
    args = parser.parse_args()

    wb = build_workbook()
    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    save_workbook(wb, args.out, force=args.force)


if __name__ == "__main__":
    main()
