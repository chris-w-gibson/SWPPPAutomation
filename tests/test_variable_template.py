import os
import sys
from openpyxl import load_workbook

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from tools.generate_variable_template import (
    PROJECT_HEADERS,
    FREQ_LIST,
    build_workbook,
    save_workbook,
)


def test_template_generation(tmp_path):
    out = tmp_path / "template.xlsx"
    wb = build_workbook()
    save_workbook(wb, out, force=True)

    assert out.exists()
    wb2 = load_workbook(out)
    ws = wb2["Projects"]
    headers = [c.value for c in ws[1]]
    assert headers == PROJECT_HEADERS

    # Check data validation on Inspection_Freq
    dv_list = [dv for dv in ws.data_validations.dataValidation if dv.type == "list"]
    found = False
    for dv in dv_list:
        if dv.formula1.replace('"', '').split(',') == FREQ_LIST:
            found = True
            break
    assert found, "Inspection_Freq validation list missing"

    # Verify cell A2 unlocked while A1 locked
    assert not ws['A2'].protection.locked
    assert ws['A1'].protection.locked
